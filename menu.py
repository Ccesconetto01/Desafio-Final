from funcoes import menu
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Gestor de Gastos Pessoal"

ws["A1"] = "Tipo"
ws["B1"] = "Categoria"
ws["C1"] = "Mes"
ws["D1"] = "Ano"
ws["E1"] = "Valor"
ws["F1"] = "Descricao"

menu(ws,wb)