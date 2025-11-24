# funcoes.py
import os
import shutil
import sqlite3
import hashlib
import hmac
import binascii
from datetime import datetime
import logging
from logging.handlers import RotatingFileHandler
from openpyxl import load_workbook, Workbook
from collections import Counter, defaultdict
import matplotlib.pyplot as plt

# -------------------------
# Config e logging (exporta logger)
# -------------------------
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)

# arquivo principal em txt conforme solicitado
RUN_LOG = os.path.join(LOG_DIR, "gestor_financeiro.txt")

logger = logging.getLogger("gestor_financeiro")
if not logger.handlers:
    logger.setLevel(logging.DEBUG)
    # RotatingFileHandler para evitar arquivos gigantes
    file_handler = RotatingFileHandler(RUN_LOG, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_formatter = logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s", "%Y-%m-%dT%H:%M:%S")
    file_handler.setFormatter(file_formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)
    console_handler.setFormatter(file_formatter)
    logger.addHandler(console_handler)

# cabeçalho inicial do log
logger.info(f"==== INICIANDO SISTEMA ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ====")

# -------------------------
# Auth (SQLite + PBKDF2)
# -------------------------
AUTH_DB = "auth.db"
PBKDF2_ITERATIONS = 200_000
HASH_NAME = "sha256"
SALT_BYTES = 16

def get_conn():
    return sqlite3.connect(AUTH_DB, timeout=5, check_same_thread=False)

def init_auth_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        iterations INTEGER NOT NULL,
        created_at TEXT NOT NULL
    )
    """)
    conn.commit()
    conn.close()
    logger.debug("[AUTH] Banco de auth inicializado (auth.db)")

def _hash_password(password: str, salt: bytes, iterations: int = PBKDF2_ITERATIONS) -> bytes:
    return hashlib.pbkdf2_hmac(HASH_NAME, password.encode('utf-8'), salt, iterations)

def create_user(username: str, password: str) -> bool:
    username = username.strip()
    if not username or not password:
        raise ValueError("Usuário e senha não podem ser vazios")
    salt = os.urandom(SALT_BYTES)
    iterations = PBKDF2_ITERATIONS
    pwd_hash = _hash_password(password, salt, iterations)
    salt_hex = binascii.hexlify(salt).decode('ascii')
    hash_hex = binascii.hexlify(pwd_hash).decode('ascii')
    created_at = datetime.utcnow().isoformat()
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("INSERT INTO users (username, password_hash, salt, iterations, created_at) VALUES (?, ?, ?, ?, ?)",
                    (username, hash_hex, salt_hex, iterations, created_at))
        conn.commit()
        conn.close()
        logger.info(f"[CADASTRO] novo_usuario={username}")
        return True
    except sqlite3.IntegrityError:
        logger.warning(f"[CADASTRO_FALHOU] usuario_existente={username}")
        return False
    except Exception:
        logger.exception("[CADASTRO_ERRO]")
        return False

def verify_user(username: str, password: str) -> bool:
    username = username.strip()
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT password_hash, salt, iterations FROM users WHERE username = ?", (username,))
        row = cur.fetchone()
        conn.close()
    except Exception:
        logger.exception("[AUTH_DB_ERROR]")
        return False
    if not row:
        logger.info(f"[LOGIN_FALHOU] usuario_nao_encontrado={username}")
        return False
    stored_hash_hex, salt_hex, iterations = row
    salt = binascii.unhexlify(salt_hex.encode('ascii'))
    calc_hash = _hash_password(password, salt, int(iterations))
    match = hmac.compare_digest(binascii.unhexlify(stored_hash_hex), calc_hash)
    if match:
        logger.info(f"[LOGIN] usuario={username}")
    else:
        logger.warning(f"[LOGIN_FALHOU] usuario={username} | senha_incorreta")
    return match

# initialize auth DB on import (idempotent)
init_auth_db()

# -------------------------
# Excel storage (com coluna Usuario)
# -------------------------
EXCEL_PATH = "Gestão.xlsx"

def ensure_excel_with_user_col(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Tipo", "Categoria", "Mês", "Ano", "Valor", "Descrição", "Usuario"])
        wb.save(path)
        logger.info(f"[EXCEL] criado '{path}' com cabeçalho (incluindo 'Usuario').")
        return
    wb = load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    if len(header) < 7 or (len(header) >= 7 and header[6] != "Usuario"):
        new_header = header[:]
        base = ["Tipo", "Categoria", "Mês", "Ano", "Valor", "Descrição"]
        for i, name in enumerate(base):
            if i >= len(new_header) or not new_header[i]:
                if i < len(new_header):
                    new_header[i] = name
                else:
                    new_header.append(name)
        if len(new_header) < 7:
            new_header.append("Usuario")
        else:
            new_header[6] = "Usuario"
        for col_idx, val in enumerate(new_header, start=1):
            ws.cell(row=1, column=col_idx, value=val)
        wb.save(path)
        logger.info(f"[EXCEL] '{path}' atualizado: cabeçalho ajustado para incluir 'Usuario'.")

ensure_excel_with_user_col(EXCEL_PATH)
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# -------------------------
# Core operations (Excel-backed)
# -------------------------
def adicionar_transacao(tipo, categoria, mes, ano, valor, descricao, user=None):
    if user is None:
        logger.warning("[ADICIONAR_NEGADO] sem usuario autenticado")
        raise PermissionError("Faça login para adicionar transações.")
    try:
        linha = ws.max_row + 1
        ws.cell(row=linha, column=1, value=tipo)
        ws.cell(row=linha, column=2, value=categoria)
        ws.cell(row=linha, column=3, value=mes)
        ws.cell(row=linha, column=4, value=ano)
        ws.cell(row=linha, column=5, value=valor)
        ws.cell(row=linha, column=6, value=descricao)
        ws.cell(row=linha, column=7, value=user)
        wb.save(EXCEL_PATH)
        # Log formatado e limpo
        try:
            mes_i = int(mes)
            ano_i = int(ano)
        except Exception:
            mes_i = mes
            ano_i = ano
        logger.info(f"[ADICIONAR] user={user} | {tipo} | {categoria} | R${float(valor):.2f} | {mes_i:02d}/{ano_i} | {descricao}")
        return linha
    except Exception:
        logger.exception("[ADICIONAR_ERRO]")
        raise

def remover_transacao(numero, user=None):
    if user is None:
        logger.warning("[REMOVER_NEGADO] sem usuario autenticado")
        raise PermissionError("Faça login para remover transações.")
    try:
        linha_real = numero + 1
        if 2 <= linha_real <= ws.max_row:
            owner = ws.cell(linha_real, 7).value
            if owner is not None and owner != user:
                logger.warning(f"[REMOVER_NEGADO_OWNER] user={user} tenta remover item de {owner} | transacao={numero}")
                return False
            tipo = ws.cell(linha_real, 1).value
            categoria = ws.cell(linha_real, 2).value
            mes = ws.cell(linha_real, 3).value
            ano = ws.cell(linha_real, 4).value
            valor = ws.cell(linha_real, 5).value
            descricao = ws.cell(linha_real, 6).value

            ws.delete_rows(linha_real)
            wb.save(EXCEL_PATH)
            logger.info(f"[REMOVER] user={user} | transacao={numero} | {tipo} | {categoria} | R${float(valor):.2f} | {mes}/{ano} | {descricao}")
            return True
        else:
            logger.warning(f"[REMOVER_INVALIDO] user={user} | transacao={numero}")
            return False
    except Exception:
        logger.exception("[REMOVER_ERRO]")
        return False

def listar_categoria(categoria):
    lista = []
    for i in range(2, ws.max_row + 1):
        if ws.cell(i, 2).value == categoria:
            lista.append(
                f"{i-1} | {ws.cell(i,1).value} | Mês:{ws.cell(i,3).value} | Ano:{ws.cell(i,4).value} | "
                f"R${ws.cell(i,5).value} | {ws.cell(i,6).value} | {ws.cell(i,7).value}"
            )
    mensagem = "\n".join(lista) if lista else "Nenhuma transação encontrada."
    logger.info(f"[LISTAR_CATEGORIA] categoria={categoria} | resultados={len(lista)}")
    return mensagem

def listar_periodo(mi, ai, mf, af):
    lista = []
    for i in range(2, ws.max_row + 1):
        try:
            mes = int(ws.cell(i, 3).value)
            ano = int(ws.cell(i, 4).value)
        except Exception:
            continue
        if (ai, mi) <= (ano, mes) <= (af, mf):
            lista.append(
                f"{i-1} | {ws.cell(i,1).value} | {ws.cell(i,2).value} | R${ws.cell(i,5).value} | {ws.cell(i,6).value} | {ws.cell(i,7).value}"
            )
    mensagem = "\n".join(lista) if lista else "Nenhuma transação no período."
    logger.info(f"[LISTAR_PERIODO] {ai:04d}-{mi:02d} -> {af:04d}-{mf:02d} | resultados={len(lista)}")
    return mensagem

def ver_saldo(username: str = None):
    """
    Se username for None -> calcula saldo global (todas transações).
    Se username for uma string -> calcula apenas as transações pertencentes a esse usuário.
    Retorna string formatada (Entradas, Saídas, Saldo).
    """
    entradas = 0.0
    saidas = 0.0
    for i in range(2, ws.max_row + 1):
        try:
            tipo = ws.cell(i, 1).value
            valor_cell = ws.cell(i, 5).value
            valor = float(valor_cell) if valor_cell not in (None, "") else 0.0
        except Exception:
            continue

        # Se username foi especificado, cheque a coluna Usuario (7)
        if username is not None:
            owner = ws.cell(i, 7).value
            if owner != username:
                continue

        if tipo == "Entrada":
            entradas += valor
        else:
            saidas += valor

    saldo = entradas - saidas
    if username:
        logger.info(f"[SALDO] user={username} | entradas={entradas:.2f} | saidas={saidas:.2f} | saldo={saldo:.2f}")
    else:
        logger.info(f"[SALDO] global | entradas={entradas:.2f} | saidas={saidas:.2f} | saldo={saldo:.2f}")

    return f"Entradas: R${entradas:.2f}\nSaídas: R${saidas:.2f}\nSaldo Final: R${saldo:.2f}"

# -------------------------
# Dados para gráficos
# -------------------------
def dados_categoria():
    categorias = []
    for i in range(2, ws.max_row + 1):
        tipo = ws.cell(i, 1).value
        categoria = ws.cell(i, 2).value
        if tipo == "Saida" and categoria is not None:
            categorias.append(str(categoria))
    cont = Counter(categorias)
    labels = list(cont.keys())
    sizes = list(cont.values())
    return labels, sizes

def dados_mensal_crescimento():
    soma = defaultdict(float)
    for i in range(2, ws.max_row + 1):
        try:
            mes = int(ws.cell(i, 3).value)
            ano = int(ws.cell(i, 4).value)
            valor = float(ws.cell(i, 5).value)
        except Exception:
            continue
        chave = (ano, mes)
        soma[chave] += valor
    pares = sorted(soma.items(), key=lambda kv: (kv[0][0], kv[0][1]))
    labels = [f"{ano:04d}-{mes:02d}" for (ano, mes), _ in pares]
    valores = [v for _, v in pares]
    cumulativo = []
    s = 0.0
    for v in valores:
        s += v
        cumulativo.append(s)
    return labels, valores, cumulativo

# Plotting helpers (opens native Matplotlib windows)
_cat_fig = None
_cat_ax = None
_month_fig = None
_month_ax = None

def open_pie_saida():
    global _cat_fig, _cat_ax
    labels, sizes = dados_categoria()
    if not labels:
        logger.info("[GRAFICO] sem dados de Saida")
        return None
    _cat_fig, _cat_ax = plt.subplots(figsize=(6,6))
    try:
        _cat_fig.canvas.manager.set_window_title("Saídas por Categoria")
    except Exception:
        pass
    _cat_ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
    _cat_ax.axis('equal')
    _cat_ax.set_title("Saídas por categoria (apenas gastos)")
    plt.show(block=False)
    return _cat_fig

def open_month_plot():
    global _month_fig, _month_ax
    labels, valores, cumulativo = dados_mensal_crescimento()
    if not labels:
        logger.info("[GRAFICO] mensal: sem dados")
        return None
    _month_fig, _month_ax = plt.subplots(figsize=(8,4))
    try:
        _month_fig.canvas.manager.set_window_title("Valor das transações mês a mês")
    except Exception:
        pass
    _month_ax.plot(labels, valores, marker='o', label='Total mensal')
    _month_ax.plot(labels, cumulativo, marker='o', linestyle='--', label='Cumulativo')
    _month_ax.set_xlabel("Mês")
    _month_ax.set_ylabel("Valor (R$)")
    _month_ax.set_title("Valor das transações mês a mês (e cumulativo)")
    _month_ax.tick_params(axis='x', rotation=45)
    _month_ax.legend()
    plt.tight_layout()
    plt.show(block=False)
    return _month_fig

# -------------------------
# Shutdown/backup helper
# -------------------------
def finalize_and_backup_logs():
    # registra encerramento
    logger.info(f"==== ENCERRANDO SISTEMA ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ====")
    handlers = logger.handlers[:]
    for h in handlers:
        try:
            h.flush()
            h.close()
        except Exception:
            pass
        logger.removeHandler(h)
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(LOG_DIR, f"gestor_financeiro_{ts}.txt")
        if os.path.exists(RUN_LOG):
            shutil.copy2(RUN_LOG, dest)
        else:
            with open(dest, "w", encoding="utf-8") as f:
                f.write(f"No run log found. Created empty final log at {ts}\n")
        return dest
    except Exception:
        # não usar logger aqui porque já fechamos handlers; ainda assim tentamos relatar
        return None
