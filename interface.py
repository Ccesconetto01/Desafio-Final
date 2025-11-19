import tkinter as tk
from openpyxl import load_workbook

wb = load_workbook("Gestão.xlsx")
ws = wb.active

def adicionar_transacao(tipo, categoria, mes, ano, valor, descricao):
    linha = ws.max_row + 1
    ws.cell(row=linha, column=1, value=tipo)
    ws.cell(row=linha, column=2, value=categoria)
    ws.cell(row=linha, column=3, value=mes)
    ws.cell(row=linha, column=4, value=ano)
    ws.cell(row=linha, column=5, value=valor)
    ws.cell(row=linha, column=6, value=descricao)
    wb.save("Gestão.xlsx")


def remover_transacao(numero):
    linha_real = numero + 1
    if 2 <= linha_real <= ws.max_row:
        ws.delete_rows(linha_real)
        wb.save("Gestão.xlsx")
        return True
    return False


def listar_categoria(categoria):
    lista = []
    for i in range(2, ws.max_row + 1):
        if ws.cell(i, 2).value == categoria:
            lista.append(
                f"{ws.cell(i,1).value} | Mês:{ws.cell(i,3).value} | Ano:{ws.cell(i,4).value} | "
                f"R${ws.cell(i,5).value} | {ws.cell(i,6).value}"
            )
    return "\n".join(lista) if lista else "Nenhuma transação encontrada."


def listar_periodo(mi, ai, mf, af):
    lista = []
    for i in range(2, ws.max_row + 1):
        mes = int(ws.cell(i, 3).value)
        ano = int(ws.cell(i, 4).value)
        if (ai, mi) <= (ano, mes) <= (af, mf):
            lista.append(
                f"{ws.cell(i,1).value} | {ws.cell(i,2).value} | R${ws.cell(i,5).value} | {ws.cell(i,6).value}"
            )
    return "\n".join(lista) if lista else "Nenhuma transação no período."


def ver_saldo():
    entradas = 0
    saidas = 0
    for i in range(2, ws.max_row + 1):
        tipo = ws.cell(i, 1).value
        valor = float(ws.cell(i, 5).value)

        if tipo == "Entrada":
            entradas += valor
        else:
            saidas += valor

    return f"Entradas: R${entradas:.2f}\nSaídas: R${saidas:.2f}\nSaldo Final: R${entradas - saidas:.2f}"


janela = tk.Tk()
janela.title("Gestor Financeiro")
janela.geometry("500x450")

container = tk.Frame(janela)
container.pack(fill="both", expand=True)

telas = {}

def criar_tela(nome):
    frame = tk.Frame(container)
    frame.place(relwidth=1, relheight=1)
    telas[nome] = frame

for nome in ["menu", "add", "remover", "categoria", "periodo", "saldo"]:
    criar_tela(nome)


def mostrar_tela(nome):
    telas[nome].tkraise()

menu = telas["menu"]

tk.Label(menu, text="MENU PRINCIPAL", font=("Arial", 16)).pack(pady=20)

tk.Button(menu, text="Adicionar transação", width=30, command=lambda: mostrar_tela("add")).pack(pady=5)
tk.Button(menu, text="Remover transação", width=30, command=lambda: mostrar_tela("remover")).pack(pady=5)
tk.Button(menu, text="Listar por categoria", width=30, command=lambda: mostrar_tela("categoria")).pack(pady=5)
tk.Button(menu, text="Listar por período", width=30, command=lambda: mostrar_tela("periodo")).pack(pady=5)
tk.Button(menu, text="Ver saldo", width=30, command=lambda: mostrar_tela("saldo")).pack(pady=5)

tk.Button(menu, text="Sair", width=30, command=janela.quit).pack(pady=20)

add = telas["add"]

tk.Label(add, text="ADICIONAR TRANSAÇÃO", font=("Arial", 14)).pack(pady=10)

e_tipo = tk.Entry(add); tk.Label(add, text="Tipo (Entrada/Saida):").pack(); e_tipo.pack()
e_cat = tk.Entry(add); tk.Label(add, text="Categoria:").pack(); e_cat.pack()
e_mes = tk.Entry(add); tk.Label(add, text="Mês:").pack(); e_mes.pack()
e_ano = tk.Entry(add); tk.Label(add, text="Ano:").pack(); e_ano.pack()
e_val = tk.Entry(add); tk.Label(add, text="Valor:").pack(); e_val.pack()
e_desc = tk.Entry(add); tk.Label(add, text="Descrição:").pack(); e_desc.pack()

msg_add = tk.Label(add, text=""); msg_add.pack(pady=10)

def confirmar_add():
    try:
        adicionar_transacao(
            e_tipo.get(),
            e_cat.get(),
            int(e_mes.get()),
            int(e_ano.get()),
            float(e_val.get()),
            e_desc.get()
        )
        msg_add["text"] = "Transação salva!"
    except:
        msg_add["text"] = "Erro: verifique os valores."

tk.Button(add, text="Salvar", command=confirmar_add).pack(pady=10)
tk.Button(add, text="Voltar", command=lambda: mostrar_tela("menu")).pack()

rem = telas["remover"]

tk.Label(rem, text="REMOVER TRANSAÇÃO", font=("Arial", 14)).pack(pady=10)

e_rem = tk.Entry(rem)
tk.Label(rem, text="Número da transação (igual ao Excel):").pack()
e_rem.pack()

msg_rem = tk.Label(rem, text=""); msg_rem.pack(pady=10)

def confirmar_rem():
    try:
        n = int(e_rem.get())
        if remover_transacao(n):
            msg_rem["text"] = "Removido!"
        else:
            msg_rem["text"] = "Número inválido!"
    except:
        msg_rem["text"] = "Digite um número válido."

tk.Button(rem, text="Remover", command=confirmar_rem).pack(pady=10)
tk.Button(rem, text="Voltar", command=lambda: mostrar_tela("menu")).pack()

cat = telas["categoria"]

tk.Label(cat, text="LISTAR POR CATEGORIA", font=("Arial", 14)).pack(pady=10)

e_cat2 = tk.Entry(cat)
tk.Label(cat, text="Categoria:").pack()
e_cat2.pack()

res_cat = tk.Label(cat, text="", justify="left")
res_cat.pack(pady=20)

def confirmar_cat():
    res_cat["text"] = listar_categoria(e_cat2.get())

tk.Button(cat, text="Listar", command=confirmar_cat).pack(pady=10)
tk.Button(cat, text="Voltar", command=lambda: mostrar_tela("menu")).pack()

per = telas["periodo"]

tk.Label(per, text="LISTAR POR PERÍODO", font=("Arial", 14)).pack(pady=10)

e_mi = tk.Entry(per); tk.Label(per, text="Mês inicial:").pack(); e_mi.pack()
e_ai = tk.Entry(per); tk.Label(per, text="Ano inicial:").pack(); e_ai.pack()
e_mf = tk.Entry(per); tk.Label(per, text="Mês final:").pack(); e_mf.pack()
e_af = tk.Entry(per); tk.Label(per, text="Ano final:").pack(); e_af.pack()

res_per = tk.Label(per, text="", justify="left")
res_per.pack(pady=20)

def confirmar_per():
    try:
        texto = listar_periodo(
            int(e_mi.get()), int(e_ai.get()),
            int(e_mf.get()), int(e_af.get())
        )
        res_per["text"] = texto
    except:
        res_per["text"] = "Erro: valores inválidos."

tk.Button(per, text="Listar", command=confirmar_per).pack(pady=10)
tk.Button(per, text="Voltar", command=lambda: mostrar_tela("menu")).pack()

sal = telas["saldo"]

tk.Label(sal, text="SALDO ATUAL", font=("Arial", 14)).pack(pady=20)

res_sal = tk.Label(sal, text="", justify="left")
res_sal.pack(pady=20)

def atualizar_saldo():
    res_sal["text"] = ver_saldo()

tk.Button(sal, text="Atualizar saldo", command=atualizar_saldo).pack(pady=10)
tk.Button(sal, text="Voltar", command=lambda: mostrar_tela("menu")).pack()

mostrar_tela("menu")
janela.mainloop()
